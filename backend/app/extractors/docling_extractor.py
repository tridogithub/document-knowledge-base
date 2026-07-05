"""Docling-based extractor for pdf / docx / pptx / xlsx with provenance metadata."""

import uuid
from pathlib import Path

from docling.chunking import HybridChunker
from docling.document_converter import DocumentConverter

from ..schema import Chunk

_converter: DocumentConverter | None = None
_chunker: HybridChunker | None = None


def _get_converter() -> DocumentConverter:
    global _converter
    if _converter is None:
        _converter = DocumentConverter()
    return _converter


def _get_chunker() -> HybridChunker:
    global _chunker
    if _chunker is None:
        _chunker = HybridChunker(max_tokens=512, merge_peers=True)
    return _chunker


def _sheet_names(path: Path) -> list[str]:
    try:
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True)
        names = list(wb.sheetnames)
        wb.close()
        return names
    except Exception:
        return []


def extract_docling(
    path: str | Path, *, project_id: str, file_id: str, file_name: str, file_type: str
) -> list[Chunk]:
    path = Path(path)
    doc = _get_converter().convert(path).document
    chunker = _get_chunker()

    sheets = _sheet_names(path) if file_type in {"xlsx", "xls"} else []

    chunks: list[Chunk] = []
    for i, dc in enumerate(chunker.chunk(doc)):
        pages = sorted(
            {
                prov.page_no
                for item in (dc.meta.doc_items or [])
                for prov in (item.prov or [])
                if getattr(prov, "page_no", None) is not None
            }
        )
        page = pages[0] if pages else None

        c = Chunk(
            id=str(uuid.uuid4()),
            project_id=project_id,
            file_id=file_id,
            file_path=file_name,
            file_name=file_name,
            file_type=file_type,
            text=dc.text,
            embed_text=chunker.contextualize(chunk=dc),
            chunk_index=i,
            section_path=list(dc.meta.headings or []),
        )
        if file_type == "pptx":
            c.slide = page
        elif file_type in {"xlsx", "xls"}:
            if page is not None and 1 <= page <= len(sheets):
                c.sheet = sheets[page - 1]
            elif page is not None:
                c.sheet = f"sheet {page}"
        else:
            c.page = page
        chunks.append(c)

    for i, c in enumerate(chunks):
        c.prev_chunk_id = chunks[i - 1].id if i > 0 else None
        c.next_chunk_id = chunks[i + 1].id if i < len(chunks) - 1 else None
    return chunks
